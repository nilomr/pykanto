#!/usr/bin/env python3

# Dependencies --------------------------

import src
from src.read.paths import safe_makedir
import pandas as pd
import numpy as np
import pygsheets
import os
from datetime import date, timedelta, datetime
from pathlib2 import Path
from texttable import Texttable
from textwrap import dedent
from tqdm.auto import tqdm

from colorama import init
from colorama import Fore, Back, Style
init()
init(autoreset=True)

# Paths -----------------------------------

PROJECT_PATH = Path("__file__").resolve().parents[2]  # ! Ego = notebook
DATA_PATH = PROJECT_PATH / "data"
FIELD_PATH = DATA_PATH / "resources" / "fieldwork" / str(date.today().year)
OUT_PATH = PROJECT_PATH / "resources" / "fieldwork" / str(date.today().year)
RPLOTS = PROJECT_PATH / "src" / "fieldwork" / "plot-new-boxes.R"

coords_xls = DATA_PATH / "resources" / "nestbox_position.xls"
recorded_xlsx = OUT_PATH / "already-recorded.xlsx"


class Error(Exception):
    pass


# Functions -------------------------------

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def yes_or_no(question):
    while "The answer is invalid":
        reply = str(input(question+' (y/n): ')).lower().strip()
        if reply[0] == 'y':
            return True
        if reply[0] == 'n':
            return False


# Main ------------------------------------

white = Fore.BLACK + Back.WHITE + Style.BRIGHT
blue = Fore.BLUE + Back.WHITE + Style.BRIGHT
red = Fore.RED + Back.WHITE + Style.BRIGHT
green = Fore.WHITE + Back.GREEN + Style.BRIGHT

while True:

    # Get coordinates for all nestboxes

    nestbox_coords = (pd.read_excel(coords_xls)
                      .filter(['x', 'y', 'Nestbox']))

    nestbox_coords['Nestbox'] = nestbox_coords['Nestbox'].str.upper()

    print(white + dedent("""
    Type 'enter' to enter newly deployed recorders
    Type 'update' to get an update about new nestboxes
    Type 'exit' to exit
    """))

    x = input().lower().strip()

    if x == "exit":
        break

    elif x != 'enter' and x != 'update':
        print(Fore.RED + Style.BRIGHT +
              "'" + x + "'" + " is not a valid command")
        continue

    elif x == 'enter':  # Enter new nestboxes

        enter_text = white + dedent("""
        Please enter all nestbox names separated by a single space
        e.g. SW84A EX20 C47
        """)

        print(enter_text)
        names = input().upper().split(' ')
        print(white + "You entered:", names)

        while not yes_or_no("Is this correct?"):
            print('\n' + white +
                  "Try again, you absolute idiot:" +
                  enter_text)
            names = input().split(' ')
            print(white + "You entered:" + blue + str(names))

        if len(names) == sum(nestbox_coords['Nestbox'].isin(names)):
            print(blue + "All nestbox names exist")
        else:
            print(red +
                  str(len(names) - sum(nestbox_coords['Nestbox'].isin(names))) +
                  " out of " + str(len(names)) + red +
                  " entered nestbox names do not exist"
                  )
            print('\n' + white +
                  "Try again, you absolute idiot:" +
                  enter_text)
            names = input().split(' ')
            print(white + "You entered:" + blue + str(names))

            while not yes_or_no("Is this correct?"):
                print('\n' + white +
                      "Try again, you absolute idiot:" +
                      enter_text)
                names = input().split(' ')
                print(white + "You entered:" + blue + str(names))

        # Enter date
        if not yes_or_no(white + "Is " + str(date.today()) +
                         " the date when you deployed these recorders?"):
            print(white + "Enter the correct date in the same format")
            day = input()
            day = datetime.strptime(day, '%Y-%m-%d').date()
        else:
            day = date.today()

        # Get coordinates, add date added and append
        new_boxes = nestbox_coords.query("Nestbox in @names")[
            ['Nestbox', 'x', 'y']]

        new_boxes['Deployed'] = str(day)
        new_boxes['Move_by'] = str(day + timedelta(days=3))

        append_df_to_excel(recorded_xlsx, new_boxes, index=False)

        print(green + "Done. You can check all added nestboxes at "
              + str(recorded_xlsx))

        continue

    elif x == 'update':  # get nestboxes to be visited

        # Auth key (keep private)
        gc = pygsheets.authorize(service_file=str(
            PROJECT_PATH / "private" / "client_secret.json"))

        # Download and append personal sheets

        workerdict = {  # ! Change every year
            "freddy": '19YSkUwkW6GyGWsk3k8Mh56OAu04q_uLubd6S7Q8CH8g',
            "julia": '1WWBT8mWVECd_lg6VJRzGHLuL78KMKojW-DCk1sjUqJU',
            "keith": '1CavBscnAw1SppJNUQ_RzZ7wGeGiSbbXh6i6M0n2MIIc',
            "charlotte": '1I1WXw55BckjETRZIb2MtlVBM7AP0ik7RV6N9O8Vhdug',
            "richard": '1G6PzcgmQ0OZF_-uFWnbCSIbZf_H4eZ9fTYnfWpYNC8Q',
            "sam": '1G92eun7KIAPVkMsDDV4_aGd1oej0jjEygRECb0sG_io',
            "samin": '1ovtFmPd5pvQCnyhO4_BFMlJdmcunpPhgd5vfChlzp70'
        }

        which_greati = pd.DataFrame(columns=['Nestbox', 'Owner'])

        for worker, googlekey in tqdm(workerdict.items(),
                                      desc="{Downloading field worker data}",
                                      position=0, leave=True):

            worker = gc.open_by_key(googlekey)[0].get_as_df(has_header=False)

            worker = (worker
                      .rename(columns=worker.iloc[0])
                      .drop(worker.index[0]))

            if '' in worker.columns:
                worker = worker.drop([''], axis=1)

            worker = (worker
                      .query("Nestbox == Nestbox")
                      .query('Species == "g" or Species == "G" or Species == "sp=g"')
                      .filter(['Nestbox', 'Owner', 'weigh eggs'])
                      .rename(columns={'weigh eggs': 'Eggs'})
                      .replace('', 'no'))

            which_greati = which_greati.append(worker)

        # Add coordinates & date added.
        # save pickle for the record

        picklename = FIELD_PATH / (str(
            str('allrounds')
            + '_'
            + str(pd.Timestamp('today', tz="UTC").strftime('%Y%m%d'))
            + '.pkl'))

        if len(which_greati) == 0:
            raise Error("There are no GRETI nestboxes")
        else:
            which_greati = pd.merge(
                which_greati, nestbox_coords, on=['Nestbox'])
            which_greati['Added'] = (str(pd.Timestamp('today', tz="UTC")
                                         .strftime('%Y-%m-%d')))
            which_greati.to_pickle(str(picklename))

        # Check which nestboxes have already been recorded

        already_recorded = pd.read_excel(recorded_xlsx).filter(['Nestbox'])

        diff_df = (which_greati.merge(already_recorded,
                                      on=['Nestbox'],
                                      indicator=True,
                                      how='outer')
                   .query('_merge != "both"')
                   .drop(['_merge'], 1)
                   .dropna(thresh=2)
                   )

        diff_df["Added"] = pd.to_datetime(diff_df["Added"])
        diff_df = diff_df.sort_values(by="Added")

        print(Fore.BLACK + Back.WHITE + dedent("""
        These are the great tit nextboxes that you haven't recorded at yet: 
        """))
        print(Fore.BLACK + Back.WHITE +
              str(diff_df.drop(['x', 'y'], 1).to_markdown()))

        print(white + dedent("""
        Type 'plot' to save a plot of all great tit nestboxes that have not been recorded
        Type 'menu' to go back to the main selector
        Type 'exit' to exit
        """))

        option = input().lower().strip()

        if option == "exit":
            break
        elif option == "plot":  # Plot in R
            diff_df.to_csv(str(OUT_PATH / 'toberecorded.csv'), index=False)
            import subprocess
            subprocess.check_call(['Rscript', str(RPLOTS)], shell=False)
            print(green + "Done. You can check your plot at "
              + str(OUT_PATH))

        elif option == "menu":
            continue
        else:
            print(Fore.RED + Style.BRIGHT +
                  "'" + option + "'" + " is not a valid command")

    # ! NOW take out all that have been marked as recorded and create a list of all
    # ! that have to be, in chronological order.

        # ! Make changes so that the program compares all great tit nestboxes #
        # ! in the spreadsheets with all the nestboxes that have already been recorded
        # ! and returns those only- so that i get not just the new ones but an incremental list
        # ! always in the same order from which i can 'tick them off'.
        # ! also make it so that:
        # ! I get 10 new nestboxes that were marked as sp=g the earliest not including those
        # ! in the 'already recorded list. for the gps points, lists and maps, include the 10 nestboxes
        # ! added to the list 3 days ago, and do so in a different colour. note: this needs to be the manually
        # ! confirmed list, since changes will happen in the field. Also add way to label nestboxes as blue tit if
        # ! i see thats the case in the field
        # !
        # !
