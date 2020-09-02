#!/usr/bin/env python3

# Dependencies --------------------------

import src
from src.read.paths import safe_makedir
import pandas as pd
import numpy as np
import pygsheets
import os
import subprocess
from pyproj import Proj, transform
from datetime import date, timedelta, datetime
from pathlib2 import Path
from texttable import Texttable
from textwrap import dedent
from tqdm.auto import tqdm

from colorama import init
from colorama import Fore, Back, Style

init()
init(autoreset=True)

# Paths -----------------------------------

PROJECT_PATH = Path("__file__").resolve().parents[0]  # ! Ego = notebook
DATA_PATH = PROJECT_PATH / "data"
FIELD_PATH = DATA_PATH / "resources" / "fieldwork" / str(date.today().year)
OUT_PATH = PROJECT_PATH / "resources" / "fieldwork" / str(date.today().year)
GPX_PATH = OUT_PATH / "gpx-files"

RPLOTS = PROJECT_PATH / "src" / "fieldwork" / "plot-new-boxes.R"
coords_xls = DATA_PATH / "resources" / "nestbox_position.xls"
recorded_xlsx = OUT_PATH / "already-recorded.xlsx"


class Error(Exception):
    pass


# Functions -------------------------------


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def yes_or_no(question):
    while "The answer is invalid":
        reply = str(input(question + " (y/n): ")).lower().strip()
        if reply[0] == "y":
            return True
        if reply[0] == "n":
            return False


def coord_transform(row):
    return pd.Series(
        transform(Proj("epsg:27700"), Proj("epsg:4326"), row["x"], row["y"])
    )


def write_gpx(filename, newboxes, tocollect):
    """Writes .gpx file to disk, containing 
    a) all great tit nestboxes that haven't been recorded (in green),
    b) nestboxes where recorders need to be collected (in red) and
    c) nestboxes that haven't been recored and have eggs (helipad symbol).

    Args:
        filename (PosixPath): path including filename and extension (.gpx) where to output file.
        newboxes (DataFrame): DataFrame containing all new boxes, with ['Nestbox', 'x', 'y'] columns.
        tocollect (DataFrame): DataFrame containing boxes from n days ago,  with ['Nestbox', 'x', 'y'] columns.
    """

    gpxfile = open(str(filename), "x")
    gpxfile.write(
        '<?xml version="1.0"?><gpx version="1.1" creator="Nilo Merino Recalde" >'
    )

    try:
        allpoints = newboxes.query('Eggs == "no"').filter(["Nestbox", "x", "y"])
        all_transformed = allpoints.apply(coord_transform, axis=1)
        allpoints_transformed = allpoints.assign(
            **{"lon": all_transformed[1], "lat": all_transformed[0]}
        ).to_dict(orient="records")

        for box in allpoints_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "poi_green"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    try:
        eggs = newboxes.query('Eggs != "no"').filter(["Nestbox", "x", "y"])
        eggs_transformed = eggs.apply(coord_transform, axis=1)
        eggs_transformed = eggs.assign(
            **{"lon": eggs_transformed[1], "lat": eggs_transformed[0]}
        ).to_dict(orient="records")

        for box in eggs_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "helipad"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    try:
        collect = tocollect.filter(["Nestbox", "x", "y"])
        coll_transformed = collect.apply(coord_transform, axis=1)
        collect_transformed = collect.assign(
            **{"lon": coll_transformed[1], "lat": coll_transformed[0]}
        ).to_dict(orient="records")

        for box in collect_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "poi_red"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    gpxfile.write("</gpx>")
    gpxfile.close()


def order(frame, var):
    if type(var) is str:
        var = [var]  # let the command take a string or list
    varlist = [w for w in frame.columns if w not in var]
    frame = frame[var + varlist]
    return frame


# Main ------------------------------------

white = Fore.BLACK + Back.WHITE + Style.BRIGHT
blue = Fore.BLUE + Back.WHITE + Style.BRIGHT
red = Fore.RED + Back.WHITE + Style.BRIGHT
green = Fore.WHITE + Back.GREEN + Style.BRIGHT

while True:

    # Get coordinates for all nestboxes

    nestbox_coords = pd.read_excel(coords_xls).filter(["x", "y", "Nestbox"])

    nestbox_coords["Nestbox"] = nestbox_coords["Nestbox"].str.upper()

    print(
        white
        + dedent(
            """
    Type 'enter' to enter newly deployed recorders
    Type 'update' to get an update asong new nestboxes
    Type 'exit' to exit
    (enter/update/exit):
    """
        )
    )

    x = input().lower().strip()

    if x == "exit":
        break

    elif x != "enter" and x != "update":
        print(Fore.RED + Style.BRIGHT + "'" + x + "'" + " is not a valid command")
        continue

    elif x == "enter":  # Enter new nestboxes and recorders
        while True:
            while True:
                print(
                    white
                    + dedent(
                        """
                Please enter all nestbox names separated by a single space:
                e.g. SW84A EX20 C47
                """
                    )
                )

                names = input().upper().strip().split(" ")

                if len(names) == sum(nestbox_coords["Nestbox"].isin(names)):
                    print(blue + "All nestbox names exist")
                    break
                else:
                    print(
                        red
                        + str(len(names) - sum(nestbox_coords["Nestbox"].isin(names)))
                        + " out of "
                        + str(len(names))
                        + red
                        + " entered nestbox names do not exist"
                    )
                    print("\n" + white + "Try again, you absolute dumbass:")
                    continue

            while True:
                print(
                    white
                    + dedent(
                        """
                Now enter the recorder numbers, also separated by spaces:
                e.g. 01 23 15
                """
                    )
                )

                recorders = input().upper().strip().split(" ")
                recorders = pd.to_numeric(recorders)

                if len(names) == len(recorders):
                    break
                else:
                    print(
                        red
                        + "The number of recorders does not match the number of nestboxes"
                    )
                    print("\n" + white + "Try again, you absolute dumbass:")
                    continue

            user_entered = dict(zip(names, recorders))

            question = (
                white
                + "You have entered: "
                + blue
                + str(user_entered)
                + white
                + dedent(
                    """
                Is this correct?"""
                )
            )
            if yes_or_no(question):
                break
            else:
                continue

            # Enter date
        if not yes_or_no(
            white
            + "Is "
            + str(date.today())
            + " the date when you deployed these recorders?"
        ):
            print(white + "Enter the correct date in the same format")
            day = input()
            day = datetime.strptime(day, "%Y-%m-%d").date()
        else:
            day = date.today()

        # Get coordinates, add date added, add recorder number and append

        new_boxes = nestbox_coords.query("Nestbox in @names")[["Nestbox", "x", "y"]]
        new_boxes["AM"] = new_boxes["Nestbox"].map(user_entered)
        new_boxes["Deployed"] = str(day)
        new_boxes["Move_by"] = str(day + timedelta(days=3))
        new_boxes = order(new_boxes, ["Nestbox", "AM"])

        append_df_to_excel(recorded_xlsx, new_boxes, index=False)

        print(
            green + "Done. You can check all added nestboxes at " + str(recorded_xlsx)
        )

        continue

    elif x == "update":  # get nestboxes to be visited

        # Auth key (keep private)
        gc = pygsheets.authorize(
            service_file=str(PROJECT_PATH / "private" / "client_secret.json")
        )

        # Download and append personal sheets

        workerdict = {  # ! Change every year
            "freddy": "1ZDNQcU5wTFn7Ac_NPm8x0jcxs-b9aMupnaH50z5qRLE",
            "julia": "1WWBT8mWVECd_lg6VJRzGHLuL78KMKojW-DCk1sjUqJU",
            "keith": "1CavBscnAw1SppJNUQ_RzZ7wGeGiSbbXh6i6M0n2MIIc",
            "charlotte": "1I1WXw55BckjETRZIb2MtlVBM7AP0ik7RV6N9O8Vhdug",
            "richard": "1G6PzcgmQ0OZF_-uFWnbCSIbZf_H4eZ9fTYnfWpYNC8Q",
            "sam": "1G92eun7KIAPVkMsDDV4_aGd1oej0jjEygRECb0sG_io",
            "samin": "1ovtFmPd5pvQCnyhO4_BFMlJdmcunpPhgd5vfChlzp70",
        }

        which_greati = pd.DataFrame(columns=["Nestbox", "Owner"])

        for worker, googlekey in tqdm(
            workerdict.items(),
            desc="{Downloading field worker data}",
            position=0,
            leave=True,
        ):
            name = worker

            worker = gc.open_by_key(googlekey)[0].get_as_df(has_header=False)

            worker = worker.rename(columns=worker.iloc[0]).drop(worker.index[0])

            if "" in worker.columns:
                worker = worker.drop([""], axis=1)

            if name == "freddy":
                worker = (
                    worker.rename(columns={"CS": "Eggs"})
                    .rename(columns={"NB": "Nestbox"})
                    .rename(columns={"SP": "Species"})
                    .query("Nestbox == Nestbox")
                    .query('Species == "g" or Species == "G" or Species == "sp=g"')
                    .filter(["Nestbox", "Eggs"])
                    .replace("", "no")
                )

                worker.insert(1, "Owner", "Freddy Hillemann")

            elif name == "sam":
                worker = (
                    worker.rename(columns={"Num eggs weighed?": "Eggs"})
                    .rename(columns={"number": "Nestbox"})
                    .query("Nestbox == Nestbox")
                    .query('Species == "g" or Species == "G" or Species == "sp=g"')
                    .filter(["Nestbox", "Eggs"])
                    .replace("", "no")
                )

                worker.insert(1, "Owner", "Sam Crofts")

            else:
                worker = (
                    worker.query("Nestbox == Nestbox")
                    .query('Species == "g" or Species == "G" or Species == "sp=g"')
                    .filter(["Nestbox", "Owner", "Clutch size"])
                    .rename(columns={"Clutch size": "Eggs"})
                    .replace("", "no")
                )

            which_greati = which_greati.append(worker)

        # Add coordinates & date added.
        # save pickle for the record

        picklename = FIELD_PATH / (
            str(
                str("allrounds")
                + "_"
                + str(pd.Timestamp("today", tz="UTC").strftime("%Y%m%d"))
                + ".pkl"
            )
        )

        if len(which_greati) == 0:
            raise Error("There are no GRETI nestboxes")
        else:
            which_greati = pd.merge(which_greati, nestbox_coords, on=["Nestbox"])
            which_greati["Added"] = str(
                pd.Timestamp("today", tz="UTC").strftime("%Y-%m-%d")
            )
            which_greati.to_pickle(str(picklename))

        # Check which nestboxes have already been recorded

        already_recorded = (
            pd.read_excel(recorded_xlsx)
            .filter(["Nestbox"])
            .query('Nestbox != "Nestbox"')
        )

        diff_df = (
            which_greati.merge(
                already_recorded, on=["Nestbox"], indicator=True, how="outer"
            )
            .query('_merge != "both"')
            .drop(["_merge"], 1)
            .dropna(thresh=2)
        )

        if {"x_x", "y_y"}.issubset(diff_df.columns):
            diff_df = diff_df.drop(["x_y", "y_y"], 1).rename(
                columns={"x_x": "x", "y_x": "y"}
            )

        diff_df = diff_df.sort_values(by="Owner")

        print(
            Fore.BLACK
            + Back.WHITE
            + dedent(
                """

        You have recorded at a total of {} nestboxes.
        There are {} new nextboxes that you haven't recorded at.
        """
            ).format(len(already_recorded), len(diff_df))
        )

        print(Fore.BLACK + Back.WHITE + str(diff_df.drop(["x", "y"], 1).to_markdown()))

        newpath = OUT_PATH / str("new_" + str(date.today()) + ".csv")
        diff_df.to_csv(newpath)
        diff_df.to_csv(str(OUT_PATH / "toberecorded.csv"), index=False)

        print(green + "This dataframe has also been saved to " + str(newpath))

        print(
            white
            + dedent(
                """
        Type 'plots' to save a plot of all great tit nestboxes that have not been recorded
        AND a gpx file with the same + nestboxes where recorders need to be collected
        Type 'menu' to go back to the main selector
        Type 'exit' to exit
        (plots/menu/exit):"""
            )
        )

        option = input().lower().strip()

        if option == "exit":
            break

        elif option == "menu":
            continue

        elif option == "plots":
            # Plot with R + ggplot2
            diff_df.to_csv(str(OUT_PATH / "toberecorded.csv"), index=False)
            subprocess.check_call(["Rscript", str(RPLOTS)], shell=False)
            print(green + "Done (1/2). You can check your plot at " + str(OUT_PATH))
            # Export gpx
            today = str(date.today())
            tomorrow = str(date.today() + timedelta(days=1))
            print(
                white
                + "Do you want the .gpx file for later today ("
                + today
                + ") or tomorrow ("
                + tomorrow
                + ")? (today/tomorrow):"
            )
            whichday = input().lower().strip()
            while True:
                if whichday == "today":
                    move_today = (
                        pd.read_excel(recorded_xlsx)
                        .query('Nestbox != "Nestbox"')
                        .query("Move_by == @today")
                    )
                    write_gpx(GPX_PATH / str(str(today) + ".gpx"), diff_df, move_today)
                    break
                elif whichday == "tomorrow":
                    move_tomorrow = (
                        pd.read_excel(recorded_xlsx)
                        .query('Nestbox != "Nestbox"')
                        .query("Move_by == @tomorrow")
                    )
                    write_gpx(
                        GPX_PATH / str(str(tomorrow) + ".gpx"), diff_df, move_tomorrow
                    )
                    break
                else:
                    print(
                        Fore.RED
                        + Style.BRIGHT
                        + "'"
                        + option
                        + "'"
                        + " is not a valid command"
                    )
                    continue

            print(green + "Done (2/2). You can find your .gpx file at " + str(GPX_PATH))

        else:
            print(
                Fore.RED + Style.BRIGHT + "'" + option + "'" + " is not a valid command"
            )


# print(Fore.BLACK + Back.WHITE + which_greati.groupby(['Owner']).size().to_markdown())

